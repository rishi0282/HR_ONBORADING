from fastmcp import FastMCP
from mcp.types import TextContent
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import os
import random
mcp = FastMCP("OnboardingServer")
EXCEL_FILE = "ob_candidates.xlsx"
def generate_unique_id():
    """Generate a unique 6-digit employee ID"""
    while True:
        new_id = random.randint(100000, 999999)
        # Check if ID already exists
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            # Check all existing IDs (column A, starting from row 2)
            existing_ids = []
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value:
                    existing_ids.append(int(cell_value))
            wb.close()
            if new_id not in existing_ids:
                return new_id
        else:
            # File doesn't exist, any ID is unique
            return new_id
def assign_device_by_role(role: str) -> str:
    """
    Assigns device based on role.
    Args:
        role (str): Job role of the candidate
    Returns:
        str: Device name
    """
    role_lower = role.lower()
    # SDE or AI Developer -> MacBook
    if "sde" in role_lower or "software" in role_lower or "developer" in role_lower or "ai" in role_lower:
        return "MacBook Pro"
    # HR -> Dell
    elif "hr" in role_lower or "human resource" in role_lower:
        return "Dell Latitude"
    # Support -> HP
    elif "support" in role_lower:
        return "HP EliteBook"
    # Default for other roles
    else:
        return "Standard Laptop"
def initialize_excel():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"
        # Set headers with all required columns
        headers = [
            "ID",
            "Name",
            "Email",
            "Contact",
            "OB Date",
            "Role",
            "Location",
            "Device Name",
            "Device Allocation Status",
            "ID Card Allocation",
            "Email Setup",
            "OB Email",
            "OB Email Status",
            "Training Email Status",
            "Training Status"
        ]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = cell.font.copy(bold=True)
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 22
        wb.save(EXCEL_FILE)
        wb.close()
@mcp.tool()
def add_candidate(
    name: str,
    email: str,
    contact: str,
    ob_date: str,
    role: str,
    location: str
) -> list[TextContent]:
    """
    Adds a new onboarding candidate to the Excel sheet with all default values.
    Device is automatically assigned based on role:
    - SDE/AI Developer/Developer → MacBook Pro
    - HR → Dell Latitude
    - Support → HP EliteBook
    - Others → Standard Laptop
    Args:
        name (str): Full name of the candidate
        email (str): Email address
        contact (str): Contact number
        ob_date (str): Onboarding date (e.g., "2024-05-15")
        role (str): Job role/position
        location (str): Work location
    Returns:
        list[TextContent]: Success message with generated ID and assigned device
    """
    try:
        # Initialize Excel if needed
        initialize_excel()
        # Generate unique ID
        employee_id = generate_unique_id()
        # Assign device based on role
        device_name = assign_device_by_role(role)
        # Load workbook
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Find next empty row
        next_row = ws.max_row + 1
        # Prepare data with default values
        data = [
            employee_id,                    # ID
            name,                           # Name
            email,                          # Email
            contact,                        # Contact
            ob_date,                        # OB Date
            role,                           # Role
            location,                       # Location
            device_name,                    # Device Name (assigned by role)
            "Pending",                      # Device Allocation Status
            "Pending",                      # ID Card Allocation
            "Pending",                      # Email Setup
            "Pending",                      # OB Email
            "Pending",                      # OB Email Status
            "Pending",                      # Training Email Status
            "NOT COMPLETED"                 # Training Status
        ]
        # Add data to Excel
        for col_num, value in enumerate(data, start=1):
            ws.cell(row=next_row, column=col_num, value=value)
        # Save and close
        wb.save(EXCEL_FILE)
        wb.close()
        return [TextContent(
            type="text",
            text=f"✅ Candidate added successfully!\n\n"
                 f"Generated Employee ID: {employee_id}\n"
                 f"Name: {name}\n"
                 f"Email: {email}\n"
                 f"Role: {role}\n"
                 f"Location: {location}\n"
                 f"Onboarding Date: {ob_date}\n"
                 f"🖥️  Assigned Device: {device_name}\n\n"
                 f"📋 Default Status Values:\n"
                 f"  • Device Allocation Status: Pending\n"
                 f"  • ID Card Allocation: Pending\n"
                 f"  • Email Setup: Pending\n"
                 f"  • OB Email: Pending\n"
                 f"  • OB Email Status: Pending\n"
                 f"  • Training Email Status: Pending\n"
                 f"  • Training Status: NOT COMPLETED"
        )]
    except Exception as e:
        return [TextContent(
            type="text",
            text=f"❌ Error adding candidate: {str(e)}"
        )]
@mcp.tool()
def get_candidate(employee_id: int) -> list[TextContent]:
    """
    Retrieves complete candidate information by employee ID.
    Args:
        employee_id (int): Employee ID to search for
    Returns:
        list[TextContent]: Complete candidate information or error message
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(
                type="text",
                text="❌ No candidates database found."
            )]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Get headers
        headers = []
        for col in range(1, ws.max_column + 1):
            headers.append(ws.cell(row=1, column=col).value)
        # Search for ID
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                result = "📋 Candidate Information:\n\n"
                for col in range(1, len(headers) + 1):
                    header = headers[col - 1]
                    value = ws.cell(row=row, column=col).value or "N/A"
                    result += f"{header}: {value}\n"
                wb.close()
                return [TextContent(type="text", text=result)]
        wb.close()
        return [TextContent(
            type="text",
            text=f"❌ No candidate found with ID: {employee_id}"
        )]
    except Exception as e:
        return [TextContent(
            type="text",
            text=f"❌ Error retrieving candidate: {str(e)}"
        )]
@mcp.tool()
def update_status(
    employee_id: int,
    field: str,
    value: str
) -> list[TextContent]:
    """
    Updates a specific status field for a candidate.
    Args:
        employee_id (int): Employee ID
        field (str): Field name to update (e.g., "Device Allocation Status")
        value (str): New value for the field
    Returns:
        list[TextContent]: Success or error message
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(
                type="text",
                text="❌ No candidates database found."
            )]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Get headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=1, column=col).value
            headers[header] = col
        # Check if field exists
        if field not in headers:
            wb.close()
            return [TextContent(
                type="text",
                text=f"❌ Field '{field}' not found. Available fields: {', '.join(headers.keys())}"
            )]
        # Find employee and update
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                col_num = headers[field]
                ws.cell(row=row, column=col_num, value=value)
                wb.save(EXCEL_FILE)
                wb.close()
                return [TextContent(
                    type="text",
                    text=f"✅ Updated successfully!\n\n"
                         f"Employee ID: {employee_id}\n"
                         f"Field: {field}\n"
                         f"New Value: {value}"
                )]
        wb.close()
        return [TextContent(
            type="text",
            text=f"❌ No candidate found with ID: {employee_id}"
        )]
    except Exception as e:
        return [TextContent(
            type="text",
            text=f"❌ Error updating status: {str(e)}"
        )]
@mcp.tool()
def list_all_candidates() -> list[TextContent]:
    """
    Lists all candidates with their current status.
    Returns:
        list[TextContent]: List of all candidates
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(
                type="text",
                text="❌ No candidates database found."
            )]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row < 2:
            wb.close()
            return [TextContent(
                type="text",
                text="📋 No candidates found in the database."
            )]
        result = "📋 All Onboarding Candidates:\n\n"
        result += f"{'ID':<10} {'Name':<20} {'Role':<20} {'Device':<20} {'Training':<15}\n"
        result += "-" * 90 + "\n"
        for row in range(2, ws.max_row + 1):
            emp_id = ws.cell(row=row, column=1).value
            name = ws.cell(row=row, column=2).value
            role = ws.cell(row=row, column=6).value
            device = ws.cell(row=row, column=8).value
            training_status = ws.cell(row=row, column=15).value
            result += f"{emp_id:<10} {name:<20} {role:<20} {device:<20} {training_status:<15}\n"
        wb.close()
        return [TextContent(type="text", text=result)]
    except Exception as e:
        return [TextContent(
            type="text",
            text=f"❌ Error listing candidates: {str(e)}"
        )]

@mcp.tool()
def update_device_status(employee_id: int) -> list[TextContent]:
    """
    Updates the device allocation status from "Pending" to "Fulfilled" for a candidate.
    Args:
        employee_id (int): Employee ID whose device status needs to be updated
    Returns:
        list[TextContent]: Success or error message
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(
                type="text",
                text="❌ No candidates database found."
            )]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Find the Device Allocation Status column (column 9)
        device_status_col = 9
        # Search for the employee ID
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                # Get current status
                current_status = ws.cell(row=row, column=device_status_col).value
                # Get candidate name for confirmation message
                candidate_name = ws.cell(row=row, column=2).value
                device_name = ws.cell(row=row, column=8).value
                # Update status to "Fulfilled"
                ws.cell(row=row, column=device_status_col, value="Fulfilled")
                # Save changes
                wb.save(EXCEL_FILE)
                wb.close()
                return [TextContent(
                    type="text",
                    text=f"✅ Device allocation status updated successfully!\n\n"
                         f"Employee ID: {employee_id}\n"
                         f"Name: {candidate_name}\n"
                         f"Device: {device_name}\n"
                         f"Previous Status: {current_status}\n"
                         f"New Status: Fulfilled"
                )]
        wb.close()
        return [TextContent(
            type="text",
            text=f"❌ No candidate found with ID: {employee_id}"
        )]
    except Exception as e:
        return [TextContent(
            type="text",
            text=f"❌ Error updating device status: {str(e)}"
        )]
    

@mcp.tool()
def update_ID_CARD_status(employee_id: int) -> list[TextContent]:
    """
    Updates the ID Card Allocation status from "Pending" to "Completed" for a candidate.
    Args:
        employee_id (int): Employee ID whose ID card status needs to be updated
    Returns:
        list[TextContent]: Success or error message
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(
                type="text",
                text="❌ No candidates database found."
            )]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # ID Card Allocation is in column 10
        id_card_col = 10
        # Search for the employee ID
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                # Get current status
                current_status = ws.cell(row=row, column=id_card_col).value
                # Get candidate name for confirmation message
                candidate_name = ws.cell(row=row, column=2).value
                # Update status to "Completed"
                ws.cell(row=row, column=id_card_col, value="Completed")
                # Save changes
                wb.save(EXCEL_FILE)
                wb.close()
                return [TextContent(
                    type="text",
                    text=f"✅ ID Card allocation status updated successfully!\n\n"
                         f"Employee ID: {employee_id}\n"
                         f"Name: {candidate_name}\n"
                         f"Previous Status: {current_status}\n"
                         f"New Status: Completed"
                )]
        wb.close()
        return [TextContent(
            type="text",
            text=f"❌ No candidate found with ID: {employee_id}"
        )]
    except Exception as e:
        return [TextContent(
            type="text",
            text=f"❌ Error updating ID card status: {str(e)}"
        )]
@mcp.tool()
def update_email_setup_status(employee_id: int) -> list[TextContent]:
    """
    Updates the Email Setup status from "Pending" to "Completed" for a candidate.
    Args:
        employee_id (int): Employee ID whose email setup status needs to be updated
    Returns:
        list[TextContent]: Success or error message
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(
                type="text",
                text="❌ No candidates database found."
            )]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Email Setup is in column 11
        email_setup_col = 11
        # Search for the employee ID
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                # Get current status
                current_status = ws.cell(row=row, column=email_setup_col).value
                # Get candidate details for confirmation message
                candidate_name = ws.cell(row=row, column=2).value
                candidate_email = ws.cell(row=row, column=3).value
                # Update status to "Completed"
                ws.cell(row=row, column=email_setup_col, value="Completed")
                # Save changes
                wb.save(EXCEL_FILE)
                wb.close()
                return [TextContent(
                    type="text",
                    text=f"✅ Email setup status updated successfully!\n\n"
                         f"Employee ID: {employee_id}\n"
                         f"Name: {candidate_name}\n"
                         f"Email: {candidate_email}\n"
                         f"Previous Status: {current_status}\n"
                         f"New Status: Completed"
                )]
        wb.close()
        return [TextContent(
            type="text",
            text=f"❌ No candidate found with ID: {employee_id}"
        )]
    except Exception as e:
        return [TextContent(
            type="text",
            text=f"❌ Error updating email setup status: {str(e)}"
        )]

if __name__ == "__main__":
    # import uvicorn
    # import os
    # # Get port from environment
    # port = int(os.environ.get("PORT", 8000))
    # app = mcp._app
    # # Run with uvicorn
    # uvicorn.run(
    #     app,  # Path to your FastMCP app
    #     host="0.0.0.0",
    #     port=port,
    #     log_level="info"
    # )
    mcp.run(transport="http",port=8000)


# if __name__ == "__main__":
#     import os
#     import sys
#     # Render provides PORT environment variable
#     port = os.environ.get("PORT", "8000")
#     print(f"🚀 Starting IT Tasks Server on port {port}", file=sys.stderr)
#     # MCP should read HOST and PORT from environment
#     os.environ["MCP_HOST"] = "0.0.0.0"
#     os.environ["MCP_PORT"] = port
#     # Run with SSE transport
#     mcp.run(transport="sse")
