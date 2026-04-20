from fastmcp import FastMCP
from mcp.types import TextContent
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openai import OpenAI
from dotenv import load_dotenv
import os
import random
from datetime import datetime, timedelta

load_dotenv()
mcp = FastMCP("OnboardingServer")
openai_client = OpenAI(api_key=os.getenv("OPENAI_API_KEY"))
EXCEL_FILE = "ob_candidates.xlsx"
def generate_unique_id():
    """Generate a unique 6-digit employee ID"""
    while True:
        new_id = random.randint(100000, 999999)
        if os.path.exists(EXCEL_FILE):
            wb = load_workbook(EXCEL_FILE)
            ws = wb.active
            existing_ids = []
            for row in range(2, ws.max_row + 1):
                cell_value = ws.cell(row=row, column=1).value
                if cell_value:
                    existing_ids.append(int(cell_value))
            wb.close()
            if new_id not in existing_ids:
                return new_id
        else:
            return new_id
def assign_device_by_role(role: str) -> str:
    """Assigns device based on role"""
    role_lower = role.lower()
    if "sde" in role_lower or "software" in role_lower or "developer" in role_lower or "ai" in role_lower:
        return "MacBook Pro"
    elif "hr" in role_lower or "human resource" in role_lower:
        return "Dell Latitude"
    elif "support" in role_lower:
        return "HP EliteBook"
    else:
        return "Standard Laptop"
def initialize_excel():
    """Create Excel file with headers if it doesn't exist"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Candidates"
        headers = [
            "ID", "Name", "Email", "Contact", "OB Date", "Role", "Location",
            "Device Name", "Device Allocation Status", "ID Card Allocation",
            "Email Setup", "OB Email", "OB Email Status", "Training Email Status", "Training Status"
        ]
        for col_num, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = cell.font.copy(bold=True)
            ws.column_dimensions[get_column_letter(col_num)].width = 22
        wb.save(EXCEL_FILE)
        wb.close()
@mcp.tool()
def add_candidate(
    name: str,
    email: str,
    contact: str,
    ob_date: str,
    role: str,
    location: str
) -> list[TextContent]:
    """
    Adds a new onboarding candidate to the Excel sheet.
    Args:
        name (str): Full name
        email (str): Email address
        contact (str): Contact number
        ob_date (str): Onboarding date (YYYY-MM-DD)
        role (str): Job role
        location (str): Work location
    Returns:
        list[TextContent]: Success message with generated ID
    """
    try:
        initialize_excel()
        employee_id = generate_unique_id()
        device_name = assign_device_by_role(role)
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        next_row = ws.max_row + 1
        data = [
            employee_id, name, email, contact, ob_date, role, location,
            device_name, "Pending", "Pending", "Pending", "Pending",
            "Pending", "Pending", "NOT COMPLETED"
        ]
        for col_num, value in enumerate(data, start=1):
            ws.cell(row=next_row, column=col_num, value=value)
        wb.save(EXCEL_FILE)
        wb.close()
        return [TextContent(
            type="text",
            text=f"✅ Candidate added!\n\n"
                 f"ID: {employee_id}\n"
                 f"Name: {name}\n"
                 f"Email: {email}\n"
                 f"Role: {role}\n"
                 f"Location: {location}\n"
                 f"Device: {device_name}\n"
                 f"OB Date: {ob_date}"
        )]
    except Exception as e:
        return [TextContent(type="text", text=f"❌ Error: {str(e)}")]
@mcp.tool()
def get_candidate(employee_id: int) -> list[TextContent]:
    """
    Retrieves candidate information by ID.
    Args:
        employee_id (int): Employee ID
    Returns:
        list[TextContent]: Candidate information
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(type="text", text="❌ No database found.")]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                result = "📋 Candidate Information:\n\n"
                for col in range(1, len(headers) + 1):
                    header = headers[col - 1]
                    value = ws.cell(row=row, column=col).value or "N/A"
                    result += f"{header}: {value}\n"
                wb.close()
                return [TextContent(type="text", text=result)]
        wb.close()
        return [TextContent(type="text", text=f"❌ ID {employee_id} not found.")]
    except Exception as e:
        return [TextContent(type="text", text=f"❌ Error: {str(e)}")]
@mcp.tool()
def generate_onboarding_email_content(employee_id: int) -> list[TextContent]:
    """
    Generates onboarding email content using LLM for a specific candidate.
    Returns subject and body separately in JSON format.
    Args:
        employee_id (int): Employee ID
    Returns:
        list[TextContent]: JSON with subject, body, and recipient email
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(type="text", text="❌ No database found.")]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        candidate_data = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                candidate_data = {
                    "name": ws.cell(row=row, column=2).value,
                    "email": ws.cell(row=row, column=3).value,
                    "ob_date": ws.cell(row=row, column=5).value,
                    "role": ws.cell(row=row, column=6).value,
                    "location": ws.cell(row=row, column=7).value,
                    "device": ws.cell(row=row, column=8).value,
                }
                break
        wb.close()
        if not candidate_data:
            return [TextContent(type="text", text=f"❌ ID {employee_id} not found.")]
        prompt = f"""
Generate a professional onboarding email for:
- Name: {candidate_data['name']}
- Role: {candidate_data['role']}
- Location: {candidate_data['location']}
- Onboarding Date: {candidate_data['ob_date']}
- Device: {candidate_data['device']}
Create:
1. A clear subject line
2. Warm congratulations message
3. Mention role, location, date, and device
4. Express excitement
Format strictly as:
SUBJECT: [subject here]
BODY: [body here]
Do not include "Best regards" or signature.
"""
        response = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {"role": "system", "content": "You are an HR email writer. Follow format strictly."},
                {"role": "user", "content": prompt}
            ],
            max_tokens=400,
            temperature=0.7
        )
        content = response.choices[0].message.content
        # Parse subject and body
        lines = content.split('\n')
        subject = ""
        body = ""
        for line in lines:
            if line.startswith("SUBJECT:"):
                subject = line.replace("SUBJECT:", "").strip()
            elif line.startswith("BODY:"):
                body_index = lines.index(line)
                body = "\n".join(lines[body_index + 1:]).strip()
                break
        # Add signature
        body += "\n\nBest regards,\nHR Team"
        import json
        result = json.dumps({
            "recipient_email": candidate_data['email'],
            "subject": subject,
            "body": body,
            "candidate_name": candidate_data['name']
        }, indent=2)
        return [TextContent(
            type="text",
            text=f"✅ Email content generated!\n\n{result}"
        )]
    except Exception as e:
        return [TextContent(type="text", text=f"❌ Error: {str(e)}")]
@mcp.tool()
def update_device_status(employee_id: int) -> list[TextContent]:
    """Updates device status to Fulfilled"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(type="text", text="❌ No database found.")]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                ws.cell(row=row, column=9, value="Fulfilled")
                wb.save(EXCEL_FILE)
                wb.close()
                return [TextContent(type="text", text=f"✅ Device status updated to Fulfilled for ID {employee_id}")]
        wb.close()
        return [TextContent(type="text", text=f"❌ ID {employee_id} not found.")]
    except Exception as e:
        return [TextContent(type="text", text=f"❌ Error: {str(e)}")]
    
@mcp.tool()
def update_ID_CARD_status(employee_id: int) -> list[TextContent]:
    """Updates ID Card status to Completed"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(type="text", text="❌ No database found.")]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                ws.cell(row=row, column=10, value="Completed")
                wb.save(EXCEL_FILE)
                wb.close()
                return [TextContent(type="text", text=f"✅ ID Card status updated to Completed for ID {employee_id}")]
        wb.close()
        return [TextContent(type="text", text=f"❌ ID {employee_id} not found.")]
    except Exception as e:
        return [TextContent(type="text", text=f"❌ Error: {str(e)}")]
    
@mcp.tool()
def update_email_setup_status(employee_id: int) -> list[TextContent]:
    """Updates Email Setup status to Completed"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(type="text", text="❌ No database found.")]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                ws.cell(row=row, column=11, value="Completed")
                wb.save(EXCEL_FILE)
                wb.close()
                return [TextContent(type="text", text=f"✅ Email Setup status updated to Completed for ID {employee_id}")]
        wb.close()
        return [TextContent(type="text", text=f"❌ ID {employee_id} not found.")]
    except Exception as e:
        return [TextContent(type="text", text=f"❌ Error: {str(e)}")]
@mcp.tool()
def list_all_candidates() -> list[TextContent]:
    """Lists all candidates"""
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(type="text", text="❌ No database found.")]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row < 2:
            wb.close()
            return [TextContent(type="text", text="📋 No candidates found.")]
        result = "📋 All Candidates:\n\n"
        result += f"{'ID':<10} {'Name':<20} {'Role':<20} {'OB Date':<12} {'Training':<15}\n"
        result += "-" * 80 + "\n"
        for row in range(2, ws.max_row + 1):
            emp_id = ws.cell(row=row, column=1).value
            name = ws.cell(row=row, column=2).value
            role = ws.cell(row=row, column=6).value
            ob_date = ws.cell(row=row, column=5).value
            training = ws.cell(row=row, column=15).value
            result += f"{emp_id:<10} {name:<20} {role:<20} {str(ob_date):<12} {training:<15}\n"
        wb.close()
        return [TextContent(type="text", text=result)]
    except Exception as e:
        return [TextContent(type="text", text=f"❌ Error: {str(e)}")]
    



def calculate_business_days(start_date, num_days):
    """
    Calculates end date by adding business days (excluding weekends).
    Args:
        start_date: Starting date
        num_days: Number of business days to add
    Returns:
        End date after adding business days
    """
    current_date = start_date
    days_added = 0
    while days_added < num_days:
        current_date += timedelta(days=1)
        # Skip weekends (5=Saturday, 6=Sunday)
        if current_date.weekday() < 5:
            days_added += 1
    return current_date
def get_training_duration(role: str) -> int:
    """
    Returns training duration in business days based on role.
    Args:
        role: Job role
    Returns:
        Number of business days for training
    """
    role_lower = role.lower()
    if "sde" in role_lower or "software" in role_lower or "developer" in role_lower:
        # Check if AI Developer specifically
        if "ai" in role_lower:
            return 55  # AI Developer: 55 days
        else:
            return 35  # SDE: 35 days
    elif "support" in role_lower:
        return 17  # Support: 17 days
    else:
        return 25  # Default: 25 days
@mcp.tool()
def generate_training_email_content(employee_id: int) -> list[TextContent]:
    """
    Generates training email content with calculated training dates.
    Training starts 12 days after onboarding.
    Duration depends on role (excluding weekends):
    - SDE: 35 business days
    - AI Developer: 55 business days
    - Support: 17 business days
    Args:
        employee_id (int): Employee ID
    Returns:
        list[TextContent]: JSON with subject, body, recipient email, and training dates
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(type="text", text="❌ No database found.")]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        candidate_data = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                candidate_data = {
                    "id": ws.cell(row=row, column=1).value,
                    "name": ws.cell(row=row, column=2).value,
                    "email": ws.cell(row=row, column=3).value,
                    "ob_date": ws.cell(row=row, column=5).value,
                    "role": ws.cell(row=row, column=6).value,
                    "location": ws.cell(row=row, column=7).value,
                }
                break
        wb.close()
        if not candidate_data:
            return [TextContent(type="text", text=f"❌ ID {employee_id} not found.")]
        # Parse onboarding date
        if isinstance(candidate_data['ob_date'], str):
            ob_date = datetime.strptime(candidate_data['ob_date'], "%Y-%m-%d")
        else:
            ob_date = candidate_data['ob_date']
        # Calculate training start date (12 days after onboarding)
        training_start = ob_date + timedelta(days=12)
        # Get training duration based on role
        training_days = get_training_duration(candidate_data['role'])
        # Calculate training end date (excluding weekends)
        training_end = calculate_business_days(training_start, training_days)
        # Format dates
        training_start_str = training_start.strftime("%B %d, %Y")  # e.g., "May 27, 2024"
        training_end_str = training_end.strftime("%B %d, %Y")
        # Create LLM prompt
        prompt = f"""
You are an HR professional writing a training invitation email.
Generate a professional training email for:
- Name: {candidate_data['name']}
- Role: {candidate_data['role']}
- Location: {candidate_data['location']}
- Training Start Date: {training_start_str}
- Training End Date: {training_end_str}
- Training Duration: {training_days} business days
The email should:
1. Welcome them and mention their role
2. Inform them about the training program
3. Clearly state training start and end dates
4. Mention the training duration
5. Emphasize importance of the training
6. Encourage questions
7. Be warm and professional
Format strictly as:
SUBJECT: [subject here]
BODY: [body here]
Do not include signature.
"""
        # Call OpenAI API
        response = openai_client.chat.completions.create(
            model="gpt-4o-mini",
            messages=[
                {
                    "role": "system",
                    "content": "You are an HR training coordinator. Follow format strictly."
                },
                {
                    "role": "user",
                    "content": prompt
                }
            ],
            max_tokens=500,
            temperature=0.7
        )
        content = response.choices[0].message.content
        # Parse subject and body
        lines = content.split('\n')
        subject = ""
        body = ""
        for line in lines:
            if line.startswith("SUBJECT:"):
                subject = line.replace("SUBJECT:", "").strip()
            elif line.startswith("BODY:"):
                body_index = lines.index(line)
                body = "\n".join(lines[body_index + 1:]).strip()
                break
        # Add signature
        body += "\n\nBest regards,\nHR Training Team"
        # Create JSON response
        import json
        result = json.dumps({
            "recipient_email": candidate_data['email'],
            "subject": subject,
            "body": body,
            "candidate_name": candidate_data['name'],
            "training_start_date": training_start_str,
            "training_end_date": training_end_str,
            "training_duration_days": training_days
        }, indent=2)
        return [TextContent(
            type="text",
            text=f"✅ Training email content generated!\n\n{result}"
        )]
    except Exception as e:
        return [TextContent(type="text", text=f"❌ Error: {str(e)}")]
@mcp.tool()
def calculate_training_dates(employee_id: int) -> list[TextContent]:
    """
    Calculates and displays training dates for a candidate.
    Useful for checking dates before sending email.
    Args:
        employee_id (int): Employee ID
    Returns:
        list[TextContent]: Training date information
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(type="text", text="❌ No database found.")]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        candidate_data = None
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                candidate_data = {
                    "name": ws.cell(row=row, column=2).value,
                    "ob_date": ws.cell(row=row, column=5).value,
                    "role": ws.cell(row=row, column=6).value,
                }
                break
        wb.close()
        if not candidate_data:
            return [TextContent(type="text", text=f"❌ ID {employee_id} not found.")]
        # Parse onboarding date
        if isinstance(candidate_data['ob_date'], str):
            ob_date = datetime.strptime(candidate_data['ob_date'], "%Y-%m-%d")
        else:
            ob_date = candidate_data['ob_date']
        # Calculate training dates
        training_start = ob_date + timedelta(days=12)
        training_days = get_training_duration(candidate_data['role'])
        training_end = calculate_business_days(training_start, training_days)
        result = f"📅 Training Date Calculation\n\n"
        result += f"Employee: {candidate_data['name']}\n"
        result += f"Role: {candidate_data['role']}\n"
        result += f"Onboarding Date: {ob_date.strftime('%B %d, %Y')}\n\n"
        result += f"Training Start: {training_start.strftime('%B %d, %Y')} (12 days after onboarding)\n"
        result += f"Training End: {training_end.strftime('%B %d, %Y')}\n"
        result += f"Duration: {training_days} business days (excluding weekends)\n"
        return [TextContent(type="text", text=result)]
    except Exception as e:
        return [TextContent(type="text", text=f"❌ Error: {str(e)}")]
    

@mcp.tool()
def update_training_status(employee_id: int) -> list[TextContent]:
    """
    Updates the training status to "COMPLETED" for a candidate.
    Args:
        employee_id (int): Employee ID whose training status needs to be updated
    Returns:
        list[TextContent]: Success or error message
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(
                type="text",
                text="❌ No candidates database found."
            )]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Training Status is in column 15
        training_status_col = 15
        # Search for the employee ID
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                # Get current status
                current_status = ws.cell(row=row, column=training_status_col).value
                # Get candidate details for confirmation
                candidate_name = ws.cell(row=row, column=2).value
                candidate_role = ws.cell(row=row, column=6).value
                # Update status to "COMPLETED"
                ws.cell(row=row, column=training_status_col, value="COMPLETED")
                # Save changes
                wb.save(EXCEL_FILE)
                wb.close()
                return [TextContent(
                    type="text",
                    text=f"✅ Training status updated successfully!\n\n"
                         f"Employee ID: {employee_id}\n"
                         f"Name: {candidate_name}\n"
                         f"Role: {candidate_role}\n"
                         f"Previous Status: {current_status}\n"
                         f"New Status: COMPLETED\n\n"
                         f"🎓 Congratulations on completing the training!"
                )]
        wb.close()
        return [TextContent(
            type="text",
            text=f"❌ No candidate found with ID: {employee_id}"
        )]
    except Exception as e:
        return [TextContent(
            type="text",
            text=f"❌ Error updating training status: {str(e)}"
        )]
@mcp.tool()
def update_training_email_status(employee_id: int) -> list[TextContent]:
    """
    Updates the training email status to "SENT" for a candidate.
    Useful for tracking if training invitation was sent.
    Args:
        employee_id (int): Employee ID
    Returns:
        list[TextContent]: Success or error message
    """
    try:
        if not os.path.exists(EXCEL_FILE):
            return [TextContent(
                type="text",
                text="❌ No candidates database found."
            )]
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        # Training Email Status is in column 14
        training_email_col = 14
        # Search for the employee ID
        for row in range(2, ws.max_row + 1):
            if ws.cell(row=row, column=1).value == employee_id:
                # Get candidate details
                candidate_name = ws.cell(row=row, column=2).value
                # Update status to "SENT"
                ws.cell(row=row, column=training_email_col, value="SENT")
                # Save changes
                wb.save(EXCEL_FILE)
                wb.close()
                return [TextContent(
                    type="text",
                    text=f"✅ Training email status updated to SENT!\n\n"
                         f"Employee ID: {employee_id}\n"
                         f"Name: {candidate_name}"
                )]
        wb.close()
        return [TextContent(
            type="text",
            text=f"❌ No candidate found with ID: {employee_id}"
        )]
    except Exception as e:
        return [TextContent(
            type="text",
            text=f"❌ Error updating training email status: {str(e)}"
        )]
    

if __name__ == "__main__":
    mcp.run(transport="sse",host="0.0.0.0",port=9000)
